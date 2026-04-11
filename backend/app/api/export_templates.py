"""
Export Templates API
Generates Excel files for Amazon bulk upload
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime
import io

from app.database import get_db
from app.models.product import Product
from app.models.seller import Seller
from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")

router = APIRouter()


def _create_workbook(products: list[dict], template_type: str) -> bytes:
    """Create Excel workbook for Amazon bulk upload"""
    wb = Workbook()
    ws = wb.active

    if template_type == "price_inventory":
        ws.title = "Price & Inventory"

        # Headers
        headers = [
            "SKU", "UPC", "EAN", "Price", "Currency", "Quantity",
            "Min Price", "Max Price", "Handling Time (Days)",
            "Fulfillment Channel", "Condition", "Weight (KG)",
            "Length (CM)", "Width (CM)", "Height (CM)",
            "Product Name"
        ]

        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", font=header_font)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data
        for row_idx, product in enumerate(products, 2):
            dims = product.get("dimensions", {}) or {}
            ws.cell(row=row_idx, column=1, value=product["sku"])
            ws.cell(row=row_idx, column=2, value=product.get("upc", ""))
            ws.cell(row=row_idx, column=3, value=product.get("ean", ""))
            ws.cell(row=row_idx, column=4, value=product["price"])
            ws.cell(row=row_idx, column=5, value=product.get("currency", "EGP"))
            ws.cell(row=row_idx, column=6, value=product["quantity"])
            ws.cell(row=row_idx, column=7, value=product.get("min_price", ""))
            ws.cell(row=row_idx, column=8, value=product.get("max_price", ""))
            ws.cell(row=row_idx, column=9, value=product.get("handling_time", 0))
            ws.cell(row=row_idx, column=10, value=product.get("fulfillment_channel", "MFN"))
            ws.cell(row=row_idx, column=11, value=product.get("condition", "New"))
            ws.cell(row=row_idx, column=12, value=product.get("weight", ""))
            ws.cell(row=row_idx, column=13, value=dims.get("length", ""))
            ws.cell(row=row_idx, column=14, value=dims.get("width", ""))
            ws.cell(row=row_idx, column=15, value=dims.get("height", ""))
            ws.cell(row=row_idx, column=16, value=product["name"])

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    elif template_type == "listing_loader":
        ws.title = "Listing Loader"

        # Headers
        headers = [
            "External ID", "External ID Type", "ASIN", "SKU",
            "Price", "Currency", "Quantity", "Condition",
            "Product Name", "Brand", "Description",
            "Manufacturer", "Model Number", "Country of Origin",
            "Weight (KG)", "Length (CM)", "Width (CM)", "Height (CM)",
            "Handling Time (Days)", "Fulfillment Channel",
            "Product Type", "Bullet Point 1"
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", font=header_font)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data
        for row_idx, product in enumerate(products, 2):
            dims = product.get("dimensions", {}) or {}
            bullets = product.get("bullet_points", []) or []
            ws.cell(row=row_idx, column=1, value=product.get("upc", product.get("ean", "")))
            ws.cell(row=row_idx, column=2, value="UPC" if product.get("upc") else "EAN" if product.get("ean") else "")
            ws.cell(row=row_idx, column=3, value=product.get("asin", ""))
            ws.cell(row=row_idx, column=4, value=product["sku"])
            ws.cell(row=row_idx, column=5, value=product["price"])
            ws.cell(row=row_idx, column=6, value=product.get("currency", "EGP"))
            ws.cell(row=row_idx, column=7, value=product["quantity"])
            ws.cell(row=row_idx, column=8, value=product.get("condition", "New"))
            ws.cell(row=row_idx, column=9, value=product["name"])
            ws.cell(row=row_idx, column=10, value=product.get("brand", ""))
            ws.cell(row=row_idx, column=11, value=product.get("description", ""))
            ws.cell(row=row_idx, column=12, value=product.get("manufacturer", ""))
            ws.cell(row=row_idx, column=13, value=product.get("model_number", ""))
            ws.cell(row=row_idx, column=14, value=product.get("country_of_origin", ""))
            ws.cell(row=row_idx, column=15, value=product.get("weight", ""))
            ws.cell(row=row_idx, column=16, value=dims.get("length", ""))
            ws.cell(row=row_idx, column=17, value=dims.get("width", ""))
            ws.cell(row=row_idx, column=18, value=dims.get("height", ""))
            ws.cell(row=row_idx, column=19, value=product.get("handling_time", 0))
            ws.cell(row=row_idx, column=20, value=product.get("fulfillment_channel", "MFN"))
            ws.cell(row=row_idx, column=21, value=product.get("product_type", ""))
            ws.cell(row=row_idx, column=22, value=bullets[0] if bullets else "")

        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@router.post("/price-inventory")
async def export_price_inventory(db: Session = Depends(get_db)):
    """
    Export products to Excel for Amazon bulk price/inventory update.
    Returns .xlsx file for download.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    products = db.query(Product).all()
    if not products:
        raise HTTPException(status_code=404, detail="No products found to export")

    products_data = []
    for p in products:
        dimensions_dict = {}
        if p.dimensions:
            try:
                import json
                dimensions_dict = json.loads(p.dimensions) if isinstance(p.dimensions, str) else p.dimensions
            except Exception:
                dimensions_dict = {}

        products_data.append({
            "sku": p.sku,
            "price": float(p.price) if p.price else 0,
            "currency": getattr(p, 'currency', 'EGP') or 'EGP',
            "quantity": p.quantity or 0,
            "min_price": "",
            "max_price": "",
            "handling_time": getattr(p, 'handling_time', 0) or 0,
            "fulfillment_channel": getattr(p, 'fulfillment_channel', 'MFN') or 'MFN',
            "condition": getattr(p, 'condition', 'New') or 'New',
            "name": p.name,
            "upc": p.upc or "",
            "ean": p.ean or "",
            "weight": float(p.weight) if p.weight else "",
            "dimensions": dimensions_dict,
        })

    excel_bytes = _create_workbook(products_data, "price_inventory")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"price_inventory_{timestamp}.xlsx"

    logger.info(f"Exported {len(products_data)} products to {filename}")

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/listing-loader")
async def export_listing_loader(db: Session = Depends(get_db)):
    """
    Export products to Excel for Amazon Listing Loader template.
    Returns .xlsx file for download.
    """
    if not OPENPYXL_AVAILABLE:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    products = db.query(Product).all()
    if not products:
        raise HTTPException(status_code=404, detail="No products found to export")

    products_data = []
    for p in products:
        dimensions_dict = {}
        if p.dimensions:
            try:
                import json
                dimensions_dict = json.loads(p.dimensions) if isinstance(p.dimensions, str) else p.dimensions
            except Exception:
                dimensions_dict = {}

        bullets = []
        if p.bullet_points:
            try:
                import json
                bullets = json.loads(p.bullet_points) if isinstance(p.bullet_points, str) else p.bullet_points
            except Exception:
                bullets = []

        products_data.append({
            "sku": p.sku,
            "price": float(p.price) if p.price else 0,
            "currency": getattr(p, 'currency', 'EGP') or 'EGP',
            "quantity": p.quantity or 0,
            "condition": getattr(p, 'condition', 'New') or 'New',
            "name": p.name,
            "brand": p.brand or "",
            "description": p.description or "",
            "upc": p.upc or "",
            "ean": p.ean or "",
            "asin": "",
            "manufacturer": p.manufacturer or "",
            "model_number": p.model_number or "",
            "country_of_origin": p.country_of_origin or "",
            "weight": float(p.weight) if p.weight else "",
            "dimensions": dimensions_dict,
            "handling_time": getattr(p, 'handling_time', 0) or 0,
            "fulfillment_channel": getattr(p, 'fulfillment_channel', 'MFN') or 'MFN',
            "product_type": p.product_type or "",
            "bullet_points": bullets,
        })

    excel_bytes = _create_workbook(products_data, "listing_loader")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"listing_loader_{timestamp}.xlsx"

    logger.info(f"Exported {len(products_data)} products to {filename}")

    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
