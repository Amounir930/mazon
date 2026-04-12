"""
Excel Listing Service
ينشئ ملفات Excel من بيانات المنتجات المحلية
صيغة متوافقة مع Amazon ListingLoader template
"""
import os
from pathlib import Path
from datetime import datetime, timezone
from typing import List, Dict, Any, Optional
from loguru import logger

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session as DbSession

from app.database import SessionLocal
from app.models.product import Product

# Export directory
EXPORT_DIR = Path(os.path.expanduser("~")) / "AppData" / "Roaming" / "CrazyLister" / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

# Amazon ListingLoader columns (simplified)
LISTING_COLUMNS = [
    "sku",
    "product_name",
    "price",
    "quantity",
    "category",
    "brand",
    "description",
    "condition",
    "fulfillment_channel",
]


def generate_listing_excel(products: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    ينشئ ملف Excel بصيغة ListingLoader من قائمة المنتجات.
    
    Args:
        products: قائمة منتجات بالحقول المطلوبة
        filename: اسم الملف (اختياري)
    
    Returns:
        مسار الملف المُنشأ
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"listing_loader_{timestamp}.xlsx"
    
    filepath = EXPORT_DIR / filename
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listings"
        
        # Headers
        headers = LISTING_COLUMNS
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row_idx, product in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=product.get("sku", ""))
            ws.cell(row=row_idx, column=2, value=product.get("product_name", ""))
            ws.cell(row=row_idx, column=3, value=float(product.get("price", 0)))
            ws.cell(row=row_idx, column=4, value=int(product.get("quantity", 0)))
            ws.cell(row=row_idx, column=5, value=product.get("category", ""))
            ws.cell(row=row_idx, column=6, value=product.get("brand", ""))
            ws.cell(row=row_idx, column=7, value=product.get("description", ""))
            ws.cell(row=row_idx, column=8, value=product.get("condition", "New"))
            ws.cell(row=row_idx, column=9, value=product.get("fulfillment_channel", "MFN"))
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        wb.save(str(filepath))
        logger.info(f"Listing Excel generated: {filepath} ({len(products)} products)")
        return str(filepath)
        
    except Exception as e:
        logger.error(f"Failed to generate listing Excel: {e}")
        raise


def generate_inventory_report(products: List[Dict[str, Any]], filename: Optional[str] = None) -> str:
    """
    ينشئ تقرير مخزون Excel من المنتجات المحلية.
    
    Args:
        products: قائمة منتجات
        filename: اسم الملف (اختياري)
    
    Returns:
        مسار الملف المُنشأ
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"inventory_report_{timestamp}.xlsx"
    
    filepath = EXPORT_DIR / filename
    
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Headers
        headers = ["SKU", "Product Name", "Price", "Quantity", "Status", "Updated At"]
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="385723", end_color="385723", fill_type="solid")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row_idx, product in enumerate(products, 2):
            ws.cell(row=row_idx, column=1, value=product.get("sku", ""))
            ws.cell(row=row_idx, column=2, value=product.get("product_name", ""))
            ws.cell(row=row_idx, column=3, value=float(product.get("price", 0)))
            ws.cell(row=row_idx, column=4, value=int(product.get("quantity", 0)))
            ws.cell(row=row_idx, column=5, value=product.get("status", "Unknown"))
            ws.cell(row=row_idx, column=6, value=product.get("updated_at", ""))
        
        # Auto-adjust
        for col_idx in range(1, len(headers) + 1):
            max_length = 10
            for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        wb.save(str(filepath))
        logger.info(f"Inventory report generated: {filepath} ({len(products)} products)")
        return str(filepath)
        
    except Exception as e:
        logger.error(f"Failed to generate inventory report: {e}")
        raise


def get_products_from_db() -> List[Dict[str, Any]]:
    """يجيب المنتجات من قاعدة البيانات"""
    db = SessionLocal()
    try:
        products = db.query(Product).all()
        result = []
        for p in products:
            result.append({
                "sku": p.sku,
                "product_name": p.name,
                "price": float(p.price) if p.price else 0,
                "quantity": int(p.quantity) if p.quantity else 0,
                "category": p.category or "",
                "brand": p.brand or "",
                "description": p.description or "",
                "condition": getattr(p, 'condition', 'New') or 'New',
                "fulfillment_channel": getattr(p, 'fulfillment_channel', 'MFN') or 'MFN',
                "status": p.status or "draft",
                "updated_at": str(p.updated_at) if p.updated_at else "",
            })
        logger.info(f"Fetched {len(result)} products from database")
        return result
    except Exception as e:
        logger.error(f"Failed to fetch products: {e}")
        return []
    finally:
        db.close()
