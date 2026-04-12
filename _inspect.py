import openpyxl
path=r"C:\Users\Dell\Desktop\learn\amazon\Data\ListingLoader.xlsm"
wb=openpyxl.load_workbook(path)
ws=wb["Template"]
print(f"Template: {ws.max_row}r x {ws.max_column}c")
for c in range(1,ws.max_column+1):
 v=ws.cell(row=4,column=c).value
 if v and str(v).strip():
  print(f"  Col{c}: {v}")
wb.close()
