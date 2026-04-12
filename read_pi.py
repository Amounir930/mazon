import openpyxl  
wb=openpyxl.load_workbook('C:/Users/Dell/Desktop/learn/amazon/Data/Flat.File.PriceInventory.eg.xlsx',data_only=True)  
for sname in wb.sheetnames:  
    ws=wb[sname]  
    print(f'### {sname} (rows={ws.max_row}, cols={ws.max_column}) ###')  
    for ri,row in enumerate(ws.iter_rows(min_row=1,max_row=min(15,ws.max_row),values_only=False),1):  
        vals=[(c.column_letter,str(c.value)[:300]) for c in row if c.value]  
        if vals:  
            print(f'  R{ri}: {vals}')  
    if ws.max_row 
        print(f'  ... total rows={ws.max_row}')  
